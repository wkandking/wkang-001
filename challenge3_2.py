#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

def combine():
    wb=load_workbook(filename='D:/MY xiazai/courses.xlsx')
    sheet_students=wb.get_sheet_by_name('students')
    sheet_time=wb.get_sheet_by_name('time')
    sheet_combine=wb.copy_worksheet(sheet_students)
    sheet_combine.title='combine'

    sheet_combine["D1"]=sheet_time["C1"].value
    a=0
    for s in sheet_combine["B"]:
        a+=1
        b=0
        for t in sheet_time["B"]:C:\Users\lenovo\PycharmProject\作业\challenge3_2.py
            b+=1
            if s.value==t.value:
                c="D"+str(a)
                d="C"+str(b)
                sheet_combine[c]=sheet_time[d].value
    wb.save(filename='D:/MY xiazai/courses1.xlsx')



def split():
    wb=load_workbook(filename='D:/MY xiazai/courses1.xlsx')
    wb_2013=Workbook()
    sheet_2013=wb_2013.active
    sheet_2013.title='2013'
    sheet_2013["A1"]='创建时间'
    sheet_2013["B1"] = '课程名称'
    sheet_2013["C1"] = '学习人数'
    sheet_2013["D1"] = '学习时间'
    wb_2014 = Workbook()
    sheet_2014 = wb_2014.active
    sheet_2014.title = '2014'
    sheet_2014["A1"] = '创建时间'
    sheet_2014["B1"] = '课程名称'
    sheet_2014["C1"] = '学习人数'
    sheet_2014["D1"] = '学习时间'
    wb_2015 = Workbook()
    sheet_2015 = wb_2015.active
    sheet_2015.title = '2015'
    sheet_2015["A1"] = '创建时间'
    sheet_2015["B1"] = '课程名称'
    sheet_2015["C1"] = '学习人数'
    sheet_2015["D1"] = '学习时间'
    wb_2016 = Workbook()
    sheet_2016 = wb_2016.active
    sheet_2016.title = '2016'
    sheet_2016["A1"] = '创建时间'
    sheet_2016["B1"] = '课程名称'
    sheet_2016["C1"] = '学习人数'
    sheet_2016["D1"] = '学习时间'
    sheet_combine=wb.get_sheet_by_name('combine')
    a=2
    b=2
    c=2
    d=2
    for i in sheet_combine:
        if str(i[0].value).split('-')[0]=='2013':
            sheet_2013["A"+str(a)]=i[0].value
            sheet_2013["B" + str(a)] = i[1].value
            sheet_2013["C" + str(a)] = i[2].value
            sheet_2013["D" + str(a)] = i[3].value
            a+=1
        elif str(i[0].value).split('-')[0]=='2014':
            sheet_2014["A"+str(b)]=i[0].value
            sheet_2014["B" + str(b)] = i[1].value
            sheet_2014["C" + str(b)] = i[2].value
            sheet_2014["D" + str(b)] = i[3].value
            b+=1
        elif str(i[0].value).split('-')[0]=='2015':
            sheet_2015["A"+str(c)]=i[0].value
            sheet_2015["B" + str(c)] = i[1].value
            sheet_2015["C" + str(c)] = i[2].value
            sheet_2015["D" + str(c)] = i[3].value
            c+=1
        elif str(i[0].value).split('-')[0]=='2016':
            sheet_2016["A"+str(d)]=i[0].value
            sheet_2016["B" + str(d)] = i[1].value
            sheet_2016["C" + str(d)] = i[2].value
            sheet_2016["D" + str(d)] = i[3].value
            d+=1


    wb_2013.save(filename="D:/MY xiazai/2013.xlsx")
    wb_2014.save(filename="D:/MY xiazai/2014.xlsx")
    wb_2015.save(filename="D:/MY xiazai/2015.xlsx")
    wb_2016.save(filename="D:/MY xiazai/2016.xlsx")

if __name__=="__main__":
    combine()
    split()